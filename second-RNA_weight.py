import time
import os
import shutil
import time
import randomstring_creation
import openpyxl


# Generate the objective function
def ecriture_fichier(RNA, OUT):
    a = str(1)
    b = str(1.2)
    length = len(RNA)
    OUT.write("Maximize \n")
    for i in range(1, length):
        for j in range(i + 1, length + 1):
            if (RNA[i - 1] == 'A') and (RNA[j - 1] == 'U'):
                x = ("+ " + a + " P(%d,%d) \n" % (i, j))
                OUT.write(x)
            if (RNA[i - 1] == 'U') and (RNA[j - 1] == 'A'):
                x = ("+ " + a + " P(%d,%d) \n" % (i, j))
                OUT.write(x)
            if (RNA[i - 1] == 'C') and (RNA[j - 1] == 'G'):
                x = ("+ " + b + " P(%d,%d) \n" % (i, j))
                OUT.write(x)
            if (RNA[i - 1] == 'G') and (RNA[j - 1] == 'C'):
                x = ("+ " + b + " P(%d,%d) \n" % (i, j))
                OUT.write("+ " + b + " P(%d,%d) \n" % (i, j))
            else:
                OUT.write("+ P(%d,%d) \n" % (i, j))

    OUT.write("such that \n")

    # The following segment generates the ILP inequalities to ensure
    # that only complementary nucleotides pair.

    for i in range(1, length):
        for j in range(i + 1, length + 1):

            if (RNA[i - 1] == 'A') and (RNA[j - 1] != 'U'):
                OUT.write("P(%d,%d) = 0 \n" % (i, j))

            if (RNA[i - 1] == 'U') and (RNA[j - 1] != 'A'):
                OUT.write("P(%d,%d) = 0 \n" % (i, j))

            if (RNA[i - 1] == 'C') and (RNA[j - 1] != 'G'):
                OUT.write("P(%d,%d) = 0 \n" % (i, j))

            if (RNA[i - 1] == 'G') and (RNA[j - 1] != 'C'):
                OUT.write("P(%d,%d) =  0 \n" % (i, j))

    # The following segment generates the ILP inequalities to ensure
    # that each position is paired to at most one other position

    for i in range(1, length + 1):
        inequality = ""

        for j in range(1, i):
            inequality = inequality + " + P(%d,%d)" % (j, i)

        for j in range(i + 1, length + 1):
            inequality = inequality + " + P(%d,%d)" % (i, j)

        inequality = inequality + ' <= 1'
        OUT.write(inequality)
        OUT.write("\n")

    # The following segment generates the ILP inequalities to ensure
    # that no pairs cross.

    for h in range(1, length - 2):
        for i in range(h + 1, length - 1):
            for j in range(i + 1, length):
                for k in range(j + 1, length + 1):
                    OUT.write("P(%d,%d) + P(%d,%d) <= 1 \n" % (h, j, i, k))

    # Write out the list of binary variables
    OUT.write("Binary \n")
    for i in range(1, length):
        for j in range(i + 1, length + 1):
            OUT.write("P(%d,%d) \n" % (i, j))

    OUT.write("end \n")

workbook = openpyxl.Workbook()

def fichier_excel(length):
    times = []
    sequence = []
    path = 'w_ARN_taille_' + str(length)
    if not os.path.exists(path):
        os.mkdir(path)
    else:
        shutil.rmtree(path)
        os.mkdir(path)

    for i in range(15):
        randomstring_creation.creation_rd_RNA(length, path)

    IN = open(path + "/randomstring", 'r')
    for RNA in IN.readlines():
        sequence.append(RNA)
        file = RNA + ".lp"
        OUT = open(path + "/" + file, 'w')
        start = time.time()
        ecriture_fichier(RNA, OUT)
        times.append(time.time() - start)
        # print(time.time() - start)

    sheet = workbook.create_sheet("Temps pour ARN de taille" + str(length), 0)

    sheet.cell(1, 1).value = "Sequence de ARN"
    sheet.cell(1, 2).value = "Temps de creation avec le poids"
    sheet.cell(1, 3).value = "Temps de resolution avec gurobi avec le poids"

    x = 2
    for i in sequence:
        sheet.cell(x, 1).value = i
        x += 1
    x = 2
    for j in times:
        sheet.cell(x, 2).value = j
        x += 1

    workbook.save('Tps_contrainte_3.xlsx')


for i in range(20, 105, 10):
    fichier_excel(i)
    print(i, "fait")
